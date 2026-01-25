#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
附录表格转换工具

将Excel文件无损转换为CSV格式，便于AI直接读取。

使用方法：
    python analyze_appendix.py                    # 转换当前目录下的 附录/ 文件夹
    python analyze_appendix.py --path ./data      # 转换指定路径
    python analyze_appendix.py --file 附录1.xlsx  # 转换单个文件

作者：数学建模 Skill
版本：2.1.0
"""

import os
import sys
import argparse
from pathlib import Path
from typing import List
import openpyxl

# 设置Windows控制台UTF-8输出
if sys.platform == 'win32':
    try:
        import io
        sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
        sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')
    except:
        pass


def sanitize_sheet_name(sheet_name: str) -> str:
    """
    清理sheet名称，使其可作为文件名
    """
    invalid_chars = ['<', '>', ':', '"', '/', '\\', '|', '?', '*']
    result = sheet_name
    for char in invalid_chars:
        result = result.replace(char, '_')
    return result


def convert_xlsx_to_csv_lossless(xlsx_path: str, output_dir: str = None) -> List[str]:
    """
    将Excel文件无损转换为CSV文件

    无损转换特点：
    - 保留所有行的数据
    - 保留所有列的数据
    - 每个sheet转换为独立的CSV文件
    - 保留单元格的原始值（包括数字、文本、日期、公式结果）
    - 使用UTF-8 with BOM编码确保兼容性
    - 保留空值

    参数：
        xlsx_path: Excel文件路径
        output_dir: 输出目录，默认为原文件所在目录

    返回：
        生成的CSV文件路径列表
    """
    if output_dir is None:
        output_dir = os.path.dirname(xlsx_path)

    base_name = Path(xlsx_path).stem
    csv_files = []

    try:
        # 使用openpyxl读取，保留原始数据
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, keep_links=False)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # 生成CSV文件名
            safe_sheet_name = sanitize_sheet_name(sheet_name)
            if len(wb.sheetnames) == 1:
                # 只有一个sheet时，直接用原文件名
                csv_filename = f"{base_name}.csv"
            else:
                # 多个sheet时，添加sheet名后缀
                csv_filename = f"{base_name}_{safe_sheet_name}.csv"

            csv_path = os.path.join(output_dir, csv_filename)

            # 收集所有行数据
            all_rows = []
            max_row = ws.max_row
            max_col = ws.max_column

            for row_idx in range(1, max_row + 1):
                row_data = []
                for col_idx in range(1, max_col + 1):
                    cell = ws.cell(row_idx, col_idx)

                    # 获取单元格值
                    value = cell.value

                    # 处理不同类型的值
                    if value is None:
                        row_data.append('')
                    elif isinstance(value, (int, float, bool)):
                        # 数字和布尔值直接转换为字符串
                        row_data.append(str(value))
                    else:
                        # 其他类型（包括字符串、日期等）转换为字符串
                        row_data.append(str(value))

                all_rows.append(row_data)

            # 写入CSV文件（UTF-8 with BOM编码）
            with open(csv_path, 'w', encoding='utf-8-sig', newline='') as f:
                for row in all_rows:
                    # 处理每行数据中的引号和逗号
                    processed_row = []
                    for cell in row:
                        # 如果单元格包含逗号、引号或换行符，需要用引号包裹
                        if ',' in str(cell) or '"' in str(cell) or '\n' in str(cell):
                            # 将已有的引号转义为两个引号
                            escaped = str(cell).replace('"', '""')
                            processed_row.append(f'"{escaped}"')
                        else:
                            processed_row.append(str(cell))
                    f.write(','.join(processed_row) + '\n')

            csv_files.append(csv_path)

        wb.close()

    except Exception as e:
        print(f"❌ 转换失败 {os.path.basename(xlsx_path)}: {str(e)}")

    return csv_files


def find_appendix_files(base_path: str = None) -> List[str]:
    """
    查找附录文件夹中的所有Excel文件

    参数：
        base_path: 基础路径，默认为当前工作目录

    返回：
        Excel文件路径列表
    """
    if base_path is None:
        base_path = os.getcwd()

    # 尝试常见的附录文件夹名称
    appendix_names = ['附录', 'appendix', 'Appendix', '附件', 'data', 'Data']
    excel_files = []

    for appendix_name in appendix_names:
        appendix_path = os.path.join(base_path, appendix_name)

        if os.path.exists(appendix_path) and os.path.isdir(appendix_path):
            print(f"✅ 找到附录文件夹: {appendix_path}")

            # 查找所有Excel文件
            for filename in sorted(os.listdir(appendix_path)):
                if filename.endswith(('.xlsx', '.xls')):
                    excel_files.append(os.path.join(appendix_path, filename))

            if excel_files:
                return excel_files

    # 如果没找到，直接在当前目录查找
    for filename in sorted(os.listdir(base_path)):
        filepath = os.path.join(base_path, filename)
        if os.path.isfile(filepath) and filename.endswith(('.xlsx', '.xls')):
            excel_files.append(filepath)

    return excel_files


def main():
    parser = argparse.ArgumentParser(
        description='将Excel文件无损转换为CSV格式',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
示例：
  python analyze_appendix.py
  python analyze_appendix.py --path ./data
  python analyze_appendix.py --file 附录1.xlsx

注意：
  - Excel文件会被自动转换为同名的CSV文件
  - 每个sheet会生成一个独立的CSV文件
  - 转换后的CSV文件使用UTF-8 with BOM编码
  - 原始Excel文件不会被修改
        """
    )

    parser.add_argument(
        '--path', '-p',
        type=str,
        default=None,
        help='指定要转换的路径（默认为当前目录）'
    )

    parser.add_argument(
        '--file', '-f',
        type=str,
        default=None,
        help='转换单个文件'
    )

    args = parser.parse_args()

    # 转换单个文件
    if args.file:
        if not os.path.exists(args.file):
            print(f"❌ 错误：文件不存在 - {args.file}")
            sys.exit(1)

        if args.file.endswith(('.xlsx', '.xls')):
            csv_files = convert_xlsx_to_csv_lossless(args.file)
            if csv_files:
                print(f"✅ 转换完成，生成CSV文件:")
                for csv_file in csv_files:
                    print(f"   - {csv_file}")
        else:
            print(f"❌ 错误：不支持的文件格式 - {args.file}")

        sys.exit(0)

    # 查找并转换所有文件
    base_path = args.path if args.path else os.getcwd()

    if not os.path.exists(base_path):
        print(f"❌ 错误：路径不存在 - {base_path}")
        sys.exit(1)

    excel_files = find_appendix_files(base_path)

    if not excel_files:
        print("❌ 未找到任何Excel文件（.xlsx, .xls）")
        sys.exit(1)

    print(f"📄 找到 {len(excel_files)} 个Excel文件")

    all_csv_files = []
    for filepath in excel_files:
        csv_files = convert_xlsx_to_csv_lossless(filepath)
        all_csv_files.extend(csv_files)

    if all_csv_files:
        print(f"\n✅ 转换完成，共生成 {len(all_csv_files)} 个CSV文件:")
        for csv_file in all_csv_files:
            print(f"   - {csv_file}")
        print(f"\n💡 提示: CSV文件可以直接使用Read工具读取")


if __name__ == '__main__':
    main()
